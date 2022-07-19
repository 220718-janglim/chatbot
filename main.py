# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import pandas as pd

data_link = "C:/Users/USER/PycharmProjects/chatbot_data.xlsx"
chatbot_data = pd.read_excel(data_link)

# rule의 데이터를 split하여 list형태로 변환 후, index값과 함께 dictionary 형태로 저장
chat_dic = {}
row = 0
for rule in chatbot_data['rule']:
    chat_dic[row] = rule.split('|')
    row += 1
print(row)

def write():
    pos = row + 1  # 엑셀 시트 상 위치 설정
    data_rule = data_question.split('|')
    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws.cell(pos, 1, data_question)  # 질문 삽입
    write_ws.cell(pos, 2, data_rule)  # 질문을 rule에 따라 변경하여 삽입
    write_ws.cell(pos, 3, data_answer)  # 답변 삽입
    xlsx = pd.read_excel(data_link)
    xlsx.to_excel(data_link, index=False)
    xlsx.to_csv(data_link, index=False)


def chat(request):
    for k, v in chat_dic.items():
        index = -1
        for word in v:
            try:
                if index == -1:
                    index = request.index(word)
                else:
                    # 이전 index 값은 현재 index값보다 이전이어야 한다.
                    if index < request.index(word, index):
                        index = request.index(word, index)
                    else:  # index 값이 이상할 경우 과감하게 break를 한다
                        index = -1
                        break
            except ValueError:
                index = -1
                break
        if index > -1:
            return chatbot_data['response'][k]
    return '무슨 말인지 모르겠어요'


while True:
    req = input('대화를 입력해보세요: ')
    if req == "exit":
        break
    if req == "write":
        data_question = input("질문을 입력해주세요: ")
        data_answer = input("답변을 입력하세요: ")
        write()

    else:
        print('jarvis : ', chat(req))


